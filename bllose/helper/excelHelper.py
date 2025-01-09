from openpyxl import load_workbook

def fetch_rows(absPath: str, withOutHead: bool = True) -> list:
    # 加载 Excel 文件 
    wb = load_workbook(filename=absPath)

    # 选择活动的工作表或通过名称选择特定的工作表
    ws = wb.active  # 默认活动的工作表
    # 或者
    ws = wb['Sheet1']  # 通过名称选择工作表

    rows = []
    # 读取单元格值
    for row in ws.iter_rows(values_only=True):
        rows.append(row)

    if withOutHead:
        rows = rows[1:]

    return rows